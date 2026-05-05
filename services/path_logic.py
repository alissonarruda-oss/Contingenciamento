import pandas as pd
from pathlib import Path
from shutil import copy
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

from services.constants import HEADERS, banco, hipo, sec, trabalhistas, outros, encerradas

def tree_search(path: Path, doc_type: str, is_root: bool = True) -> list[Path]:
    """Percorre recursivamente o diretório informado e retorna todos os arquivos da extensão especificada encontrados em subpastas (exclui o diretório raiz)."""
    docs = []
    for p in path.iterdir():
        if p.is_dir():
            docs.extend(tree_search(p, doc_type, False))
        elif p.suffix == doc_type and not is_root:
            docs.append(p)
    if not len(docs) > 0:
        return []
    else:
        return docs


def salvar_aba(lista_dfs: list[pd.Series], writer: pd.ExcelWriter, nome_aba: str, colunas_esperadas: list[str] | None = None) -> None:
    """Grava uma lista de registros como aba Excel, aplicando formatação visual: cabeçalho preto/branco, largura automática de colunas, destaque amarelo em colunas de erro e vermelho em campos críticos vazios."""
    if lista_dfs:
        df_final = pd.DataFrame(lista_dfs).drop_duplicates()
    else:
        if colunas_esperadas:
            df_final = pd.DataFrame(columns=colunas_esperadas)
        else:
            df_final = pd.DataFrame()

    df_final.to_excel(writer, sheet_name=nome_aba, index=False)
    ws = writer.sheets[nome_aba]

    fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    fundo_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fundo_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    cols_monitoradas = ["ESCRITÓRIO", "PARTE AUTORA", "PARTE RÉ", "PRODUTO"]
    colunas_monetarias = ["VALOR DA CAUSA", "VALOR DO RISCO ATUALIZADO"]

    for col_idx, col_name in enumerate(df_final.columns, start=1):
        col_letter = get_column_letter(col_idx)

        celula_cabecalho = ws.cell(row=1, column=col_idx)
        celula_cabecalho.fill = fundo_preto
        celula_cabecalho.font = fonte_branca

        tamanho_maximo = len(str(col_name))
        if not df_final.empty:
            tamanho_maximo = max(df_final[col_name].astype(str).map(len).max(), tamanho_maximo)

        ws.column_dimensions[col_letter].width = min(tamanho_maximo + 2, 70)

        nome_coluna_atual = str(col_name).upper()

        for row_idx in range(2, ws.max_row + 1):
            celula = ws.cell(row=row_idx, column=col_idx)

            if nome_coluna_atual in ["ARQUIVO_ORIGEM", "ABA_ORIGEM", "PROBLEMA"]:
                celula.fill = fundo_amarelo
            elif nome_coluna_atual in cols_monitoradas:
                if celula.value is None or str(celula.value).strip() == "" or str(celula.value).lower() == "nan":
                    celula.fill = fundo_vermelho

            if nome_coluna_atual in colunas_monetarias:
                if celula.value is not None and type(celula.value) in [int, float]:
                    celula.number_format = "#,##0.00"


def exportar_consolidado() -> None:
    """Gera o arquivo CONSOLIDADO - HIPOTECÁRIA_BANCO_SEC.xlsx copiando a base dados.xlsx e adicionando abas separadas para Banco, Hipotecária e Securitizadora (ativas e passivas)."""
    consolidado_path = "CONSOLIDADO - HIPOTECÁRIA_BANCO_SEC.xlsx"
    copy("dados.xlsx", consolidado_path)

    with pd.ExcelWriter(consolidado_path, engine="openpyxl", mode="a") as wr:
        salvar_aba(banco["ativas"], wr, "BANCO - ATIVAS", HEADERS["BH(ATIVAS)"])
        salvar_aba(banco["passivas"], wr, "BANCO - PASSIVAS", HEADERS["BH(PASSIVAS)"])
        salvar_aba(hipo["ativas"], wr, "HIPO - ATIVAS", HEADERS["BH(ATIVAS)"])
        salvar_aba(hipo["passivas"], wr, "HIPO - PASSIVAS", HEADERS["BH(PASSIVAS)"])
        salvar_aba(sec["ativas"], wr, "SEC - ATIVAS", HEADERS["SEC(ATIVAS)"])
        salvar_aba(sec["passivas"], wr, "SEC - PASSIVAS", HEADERS["SEC(PASSIVAS)"])

        workbook = wr.book
        workbook.move_sheet("DADOS", offset=len(workbook.sheetnames))


def exportar_trabalhistas() -> None:
    """Gera dois arquivos Excel de ações trabalhistas: um para Service e Promotora e outro para Banco e Hipotecária, cada um com abas por entidade."""
    with pd.ExcelWriter("TRABALHISTA_CONSOLIDADO - SERVICE e PROMOTORA.xlsx", engine="openpyxl") as wr:
        salvar_aba(trabalhistas["SERVICE"], wr, "AÇÕES TRABALHISTAS - SERVICE", HEADERS["TRABALHISTAS"])
        salvar_aba(trabalhistas["PROMOTORA"], wr, "AÇÕES TRABALHISTAS - PROMOTORA", HEADERS["TRABALHISTAS"])

    with pd.ExcelWriter("TRABALHISTA_CONSOLIDADO - BANCO E HIPO.xlsx", engine="openpyxl") as wr:
        salvar_aba(trabalhistas["BANCO"], wr, "AÇÕES TRABALHISTAS - BANCO", HEADERS["TRABALHISTAS"])
        salvar_aba(trabalhistas["HIPO"], wr, "AÇÕES TRABALHISTAS - HIPO", HEADERS["TRABALHISTAS"])


def exportar_outros() -> None:
    """Gera o arquivo VERIFICAR_OUTROS.xlsx com todos os registros que não puderam ser classificados, organizados em abas por escritório para facilitar revisão manual."""
    if not outros:
        return

    with pd.ExcelWriter("VERIFICAR_OUTROS.xlsx", engine="openpyxl") as wr:
        escritorios = set(row.get("ESCRITÓRIO", "Sem escritorio especificado") for row in outros)
        for escritorio in escritorios:
            linhas_do_escritorio = [row for row in outros if row.get("ESCRITÓRIO") == escritorio]
            nome_aba = ""
            if len(linhas_do_escritorio) > 0:
                nome_aba = str(escritorio).replace(":", "-").replace("/", "-")
                nome_aba = nome_aba[:31].strip()
            else:
                linhas_do_escritorio = [row for row in outros if row.get("ESCRITÓRIO", "") == ""]
                nome_aba = "Sem escritorio especificado"

            salvar_aba(linhas_do_escritorio, wr, nome_aba)

    print("\n⚠️ Alguns registros não foram classificados. Verifique o arquivo 'VERIFICAR_OUTROS.xlsx'")


def exportar_encerradas() -> None:
    """Gera o arquivo ENCERRADAS.xlsx com todas as ações encerradas identificadas nas planilhas de entrada, removendo duplicatas."""
    if not encerradas:
        return

    pd.DataFrame(encerradas).drop_duplicates().to_excel("ENCERRADAS.xlsx", index=False)
             