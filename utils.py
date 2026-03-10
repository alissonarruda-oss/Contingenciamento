import pandas as pd
from pathlib import Path

from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter

HEADERS = {
    "ESSENCIAIS": [
        "ESCRITÓRIO",
        "CNPJ",
        "PARTE AUTORA",
        "PARTE RÉ",
        "PROCESSO",
        "UF",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "PRODUTO",
    ],
    "BH(ATIVAS)": [
        "ESCRITÓRIO",
        "CNPJ",
        "PARTE AUTORA",
        "PARTE RÉ",
        "CPF/CNPJ RÉ",
        "NÚMERO DO PROCESSO",
        "UF",
        "COMARCA",
        "JUÍZO",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "ASSUNTO",
        "LIMINAR",
        "VALOR DA CAUSA",
        "VALOR DO RISCO ATUALIZADO",
        "PROBABILIDADE DE PERDA",
        "HONORÁRIOS",
        "DEPÓSITOS BARI",
        "DEPÓSITOS ATUALIZADOS - BARI",
        "DEPÓSITOS - AUTORA",
        "ÚLTIMO ANDAMENTO PROCESSUAL",
        "PRODUTO",
        "OBSERVAÇÃO",
    ],
    "BH(PASSIVAS)": [
        "ESCRITÓRIO",
        "CNPJ",
        "PARTE AUTORA",
        "PARTE RÉ",
        "NÚMERO DO PROCESSO",
        "UF",
        "COMARCA",
        "JUÍZO",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "ASSUNTO",
        "LIMINAR",
        "VALOR DA CAUSA",
        "VALOR DO RISCO ATUALIZADO",
        "PROBABILIDADE DE PERDA",
        "HONORÁRIOS",
        "DEPÓSITOS BARI",
        "DEPÓSITOS ATUALIZADOS - BARI",
        "DEPÓSITOS - CLIENTE",
        "ÚLTIMO ANDAMENTO PROCESSUAL",
        "PRODUTO",
        "OBSERVAÇÃO",
    ],
    "SEC(ATIVAS)": [
        "ESCRITÓRIO",
        "CNPJ",
        "PARTE AUTORA",
        "PARTE RÉ",
        "CPF/CNPJ RÉ",
        "CRI",
        "NÚMERO DO PROCESSO",
        "UF",
        "COMARCA",
        "JUÍZO",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "ASSUNTO",
        "LIMINAR",
        "VALOR DA CAUSA",
        "VALOR DO RISCO ATUALIZADO",
        "PROBABILIDADE DE PERDA",
        "HONORÁRIOS",
        "DEPÓSITOS BARI",
        "DEPÓSITOS ATUALIZADOS - BARI",
        "DEPÓSITOS - AUTORA",
        "ÚLTIMO ANDAMENTO PROCESSUAL",
        "PRODUTO",
        "OBSERVAÇÃO",
    ],
    "SEC(PASSIVAS)": [
        "ESCRITÓRIO",
        "CNPJ",
        "PARTE AUTORA",
        "PARTE RÉ",
        "CRI",
        "NÚMERO DO PROCESSO",
        "UF",
        "COMARCA",
        "JUÍZO",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "ASSUNTO",
        "LIMINAR",
        "VALOR DA CAUSA",
        "VALOR DO RISCO ATUALIZADO",
        "PROBABILIDADE DE PERDA",
        "HONORÁRIOS",
        "DEPÓSITOS BARI",
        "DEPÓSITOS ATUALIZADOS - BARI",
        "DEPÓSITOS - CLIENTE",
        "ÚLTIMO ANDAMENTO PROCESSUAL",
        "PRODUTO",
        "OBSERVAÇÃO",
    ],
    "TRABALHISTAS": [
        "ESCRITÓRIO",
        "CNPJ ESCRITÓRIO",
        "PARTE AUTORA",
        "PARTE RÉ",
        "Nº PROCESSO",
        "UF",
        "COMARCA",
        "JUÍZO",
        "DATA DA AÇÃO",
        "TIPO DE AÇÃO",
        "ASSUNTO DA AÇÃO",
        "LIMINAR ATIVA",
        "VALOR ORIGINAL DO LITÍGIO",
        "VALOR DO RISCO ATUALIZADO",
        "PROBABILIDADE DE PERDA",
        "HONORÁRIOS DE ÊXITO",
        "DEPÓSITOS RECLAMANTE",
        "DEPÓSITO RECURSO ORDINÁRIO",
        "DEPÓSITO RECURSO DE REVISTA",
        "GARANTIA DO JUÍZO",
        "DEPÓSITOS BARI",
        "ÚLTIMO ANDAMENTO PROCESSUAL",
        "PRODUTO",
        "OBSERVAÇÃO",
    ],
}

COLS = ["ESCRITÓRIO", "PARTE AUTORA", "PARTE RÉ", "NÚMERO DO PROCESSO", "PRODUTO", "VALOR DA CAUSA", "VALOR DO RISCO ATUALIZADO", "PROBABILIDADE DE PERDA"]
MONETARIO = ["VALOR DA CAUSA", "VALOR DO RISCO ATUALIZADO"]

banco: dict[str, list[pd.DataFrame]] = {"ativas": [], "passivas": []}
hipo: dict[str, list[pd.DataFrame]] = {"ativas": [], "passivas": []}
sec: dict[str, list[pd.DataFrame]] = {"ativas": [], "passivas": []}
trabalhistas: dict[str, list[pd.DataFrame]] = {"BANCO": [], "SERVICE": [], "PROMOTORA": [], "HIPO":[]}
outros: list[pd.Series] = []


# Adicione o parâmetro 'colunas_esperadas' na função
def salvar_aba(lista_dfs: list[pd.Series], writer, nome_aba, colunas_esperadas=None) -> None:

    # Se tiver dados, processa normalmente
    if lista_dfs:
        df_final = pd.DataFrame(lista_dfs).drop_duplicates()
    else:
        # Se não tiver dados, cria um DataFrame VAZIO, mas já com as colunas
        if colunas_esperadas:
            df_final = pd.DataFrame(columns=colunas_esperadas)
        else:
            df_final = pd.DataFrame()

    # O to_excel agora roda SEMPRE, garantindo que a aba exista
    df_final.to_excel(writer, sheet_name=nome_aba, index=False)
    ws = writer.sheets[nome_aba]

    fundo_preto = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    fonte_branca = Font(color="FFFFFF", bold=True)
    fundo_amarelo = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    fundo_vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    cols_monitoradas = ["ESCRITÓRIO", "PARTE AUTORA", "PARTE RÉ", "PRODUTO"]
    colunas_monetarias = ["VALOR DA CAUSA", "VALOR DO RISCO ATUALIZADO"]

    # Como o df_final terá colunas (mesmo sem dados), ele vai pintar o cabeçalho de preto!
    for col_idx, col_name in enumerate(df_final.columns, start=1):
        col_letter = get_column_letter(col_idx)

        celula_cabecalho = ws.cell(row=1, column=col_idx)
        celula_cabecalho.fill = fundo_preto
        celula_cabecalho.font = fonte_branca

        # Ajuste seguro da largura (evita erro se a coluna estiver vazia)
        tamanho_maximo = len(str(col_name))
        if not df_final.empty:
            tamanho_maximo = max(df_final[col_name].astype(str).map(len).max(), tamanho_maximo)

        ws.column_dimensions[col_letter].width = min(tamanho_maximo + 2, 70)

        nome_coluna_atual = str(col_name).upper()

        # O resto do loop de formatação das linhas continua igualzinho...
        for row_idx in range(2, ws.max_row + 1):
            celula = ws.cell(row=row_idx, column=col_idx)

            if nome_coluna_atual in ["ARQUIVO_ORIGEM", "ABA_ORIGEM", "PROBLEMA"]:
                celula.fill = fundo_amarelo
            elif nome_coluna_atual in cols_monitoradas:
                if celula.value is None or str(celula.value).strip() == "" or str(celula.value).lower() == "nan":
                    celula.fill = fundo_vermelho

            if nome_coluna_atual in colunas_monetarias:
                if celula.value is not None and type(celula.value) in [int, float]:
                    celula.number_format = "#,##0.00"


def tree_search(path: Path, doc_type: str, is_root: bool = True) -> list[Path]:
    docs = []
    for p in path.iterdir():
        if p.is_dir():
            docs.extend(tree_search(p, doc_type, False))
        elif p.suffix == doc_type and not is_root:
            docs.append(p)
    if not len(docs) > 0:
        return []
    else:
        return docs


def validador(page_headers, arr_final: list, outros: list, row: pd.Series, doc: Path, nome: str):
    if len(page_headers) > 0:
        if len(row) == len(page_headers):
            row.index = page_headers
            if all(pd.notna(row[col]) for col in COLS):
                for mon in MONETARIO:
                    try:
                        valor_str = str(row[mon]).replace("R$", "")

                        if valor_str and valor_str.lower() != "nan":
                            row[mon] = float(valor_str)

                    except ValueError:
                        row["ARQUIVO_ORIGEM"] = doc.name
                        row["ABA_ORIGEM"] = nome
                        row["PROBLEMA"] = f"Valor não monetário"
                        outros.append(row)

                        return

                arr_final.append(row)
            else:
                row["ARQUIVO_ORIGEM"] = doc.name
                row["ABA_ORIGEM"] = nome
                row["PROBLEMA"] = "Informações não preenchidas"
                outros.append(row)
        else:
            row["ARQUIVO_ORIGEM"] = doc.name
            row["ABA_ORIGEM"] = nome
            row["PROBLEMA"] = "Quantidade de colunas incorretas"
            outros.append(row)
    else:
        arr_final.append(row)
